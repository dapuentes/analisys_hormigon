# calculosh/reportes.py
import pandas as pd
from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows # No se usa directamente si _escribir_dataframe_a_hoja_v2 está bien
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import numpy as np
from datetime import datetime
import openpyxl # Para get_column_letter

# --- Helper Function para Formateo Seguro ---
def format_value(value, specifier=""):
    """Formatea un valor de forma segura, manejando strings y N/A."""
    if isinstance(value, (int, float)) and not np.isnan(value): # Si es un número válido
        return f"{value:{specifier}}"
    elif isinstance(value, str):
        try:
            # Intentar convertir a float si es un string que representa un número
            float_val = float(value)
            return f"{float_val:{specifier}}"
        except ValueError:
            # Si no se puede convertir (ej. es "N/D"), devolver el string original
            return value
    elif pd.notna(value): # Para otros tipos que no son NaN y no son string (raro aquí)
        return str(value)
    return "N/D" # Default para None, NaN

# --- Estilos (Revisados y con nombres únicos para evitar conflictos si se recarga) ---
def aplicar_estilos_reporte(ws):
    """Aplica estilos básicos y mejorados a una hoja de trabajo."""
    title_font_style = Font(bold=True, name='Calibri', size=14, color="000000")
    header_font_style = Font(bold=True, color="FFFFFF", name='Calibri', size=11)
    header_fill_style = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Estilo para celdas de datos normales
    normal_style_reporte = NamedStyle(name="normal_style_memoria_v5") # Nombre único
    normal_style_reporte.font = Font(name='Calibri', size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    normal_style_reporte.border = thin_border
    normal_style_reporte.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Estilo para encabezados de tabla
    header_style_reporte = NamedStyle(name="header_style_memoria_v5") # Nombre único
    header_style_reporte.font = header_font_style
    header_style_reporte.fill = header_fill_style
    header_style_reporte.border = thin_border
    header_style_reporte.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Estilo para celdas de parámetros (clave-valor)
    param_key_style = NamedStyle(name="param_key_style_v5")
    param_key_style.font = Font(name='Calibri', size=10, bold=True)
    param_key_style.border = thin_border
    param_key_style.alignment = Alignment(horizontal='left', vertical='center')

    param_value_style = NamedStyle(name="param_value_style_v5")
    param_value_style.font = Font(name='Calibri', size=10)
    param_value_style.border = thin_border
    param_value_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Registrar estilos si no existen para evitar errores en algunas versiones de openpyxl o recargas
    # Puede ser opcional si los nombres son siempre únicos por sesión.
    # styles_to_register = [normal_style_reporte, header_style_reporte, param_key_style, param_value_style]
    # for style in styles_to_register:
    #     if style.name not in ws.parent.named_styles:
    #         try:
    #             ws.parent.add_named_style(style)
    #         except ValueError: # El estilo ya podría existir con el mismo nombre pero diferente ID interno
    #             pass


    ws.sheet_view.showGridLines = True
    return title_font_style, header_style_reporte, normal_style_reporte, param_key_style, param_value_style

def _escribir_tabla_simple_v2(ws, data_dict, start_row, title_font, param_key_style, param_value_style, title, title_cols_merged=2):
    if title:
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = title_font
        title_cell.alignment = Alignment(vertical='center')
        if title_cols_merged > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=title_cols_merged)
        start_row += 1
    
    for key, value in data_dict.items():
        ws.cell(row=start_row, column=1, value=str(key)).style = param_key_style
        val_to_write = "N/D" # Default
        # El 'value' aquí ya viene formateado por format_value si es numérico, o como string
        if pd.notna(value): 
            val_to_write = str(value)
        ws.cell(row=start_row, column=2, value=val_to_write).style = param_value_style
        start_row += 1
    
    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 40 # Ancho para claves
    ws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 50 # Ancho para valores
    return start_row + 1

def _escribir_dataframe_a_hoja_v2(ws, df, start_row, title_font, header_style, normal_style, table_title, max_col_width=35, min_col_width=10):
    if table_title:
        title_cell = ws.cell(row=start_row, column=1, value=table_title)
        title_cell.font = title_font
        title_cell.alignment = Alignment(vertical='center')
        if df.shape[1] > 1:
             ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=df.shape[1])
        start_row += 1
        
    for c_idx, col_name in enumerate(df.columns):
        ws.cell(row=start_row, column=c_idx + 1, value=col_name).style = header_style
    current_data_row = start_row + 1
    
    for r_idx, row_data in enumerate(df.itertuples(index=False), start=current_data_row):
        for c_idx, value in enumerate(row_data):
            cell = ws.cell(row=r_idx, column=c_idx + 1)
            if pd.isna(value) or value is None or (isinstance(value, float) and np.isnan(value)):
                cell.value = "N/D"
            elif isinstance(value, (float, np.floating)): # Formatear floats en DataFrames
                 # Asumir 2-3 decimales para DataFrames, puede ser parametrizado
                cell.value = round(value, 3) if abs(value) > 1e-3 else f"{value:.2e}" 
            else:
                cell.value = value
            cell.style = normal_style
    
    for i, column_name in enumerate(df.columns):
        column_letter = openpyxl.utils.get_column_letter(i + 1)
        max_l = len(str(column_name)) 
        for r in range(len(df)):
            cell_val = df.iat[r, i]
            if pd.notna(cell_val):
                # Si el valor ya es un string (como "N/D"), o si es un número formateado
                # La longitud se toma directamente.
                max_l = max(max_l, len(str(cell_val if not isinstance(cell_val, (float,np.floating)) else round(cell_val,3) ) ) )
        
        adjusted_width = min(max((max_l + 3) * 1.15, min_col_width), max_col_width)
        ws.column_dimensions[column_letter].width = adjusted_width

    return current_data_row + len(df) + 1


def generar_memoria_excel(datos_memoria, nombre_archivo="Memoria_Calculo_Estructural.xlsx"):
    wb = Workbook()
    wb.remove(wb.active) 

    # --- 0. Hoja: Portada y Parámetros Globales ---
    ws_portada = wb.create_sheet("Portada y Parámetros")
    tf_pt, hs_pt, ns_pt, pks_pt, pvs_pt = aplicar_estilos_reporte(ws_portada)
    
    current_row_pt = 1
    ws_portada.cell(row=current_row_pt, column=1, value="MEMORIA DE CÁLCULO ESTRUCTURAL").font = Font(bold=True, size=20, name='Calibri', color="4F81BD")
    ws_portada.merge_cells(start_row=current_row_pt, start_column=1, end_row=current_row_pt, end_column=5)
    ws_portada.cell(row=current_row_pt, column=1).alignment = Alignment(horizontal='center')
    current_row_pt += 2

    info_proyecto = datos_memoria.get("info_proyecto", {})
    pg = datos_memoria.get("parametros_globales", {})
    data_portada_proyecto = {
        "Nombre del Proyecto": info_proyecto.get("nombre_proyecto", pg.get("nombre_proyecto", "N/D")),
        "Localización": info_proyecto.get("localizacion", pg.get("localizacion", "N/D")),
        "Fecha de Generación": info_proyecto.get("fecha", datetime.now().strftime("%Y-%m-%d")),
        "Normativa Principal Aplicada": info_proyecto.get("normativa_principal", "NSR-10 Colombia"),
        "Ingeniero(s) Responsable(s)": info_proyecto.get("ingenieros_responsables", "N/D")
    }
    current_row_pt = _escribir_tabla_simple_v2(ws_portada, data_portada_proyecto, current_row_pt, tf_pt, pks_pt, pvs_pt, "Información del Proyecto")

    fc_losas_vigas = pg.get('fc_losas_vigas_MPa', 21) # Default si no se encuentra
    ec_concreto_str = "N/A"
    if pd.notna(fc_losas_vigas) and isinstance(fc_losas_vigas, (int, float)) and fc_losas_vigas > 0:
        ec_concreto_str = format_value(4700 * np.sqrt(fc_losas_vigas), ".0f")

    material_data_dict = { # Renombré para claridad
        "f'c Columnas (MPa)": format_value(pg.get('fc_columnas_MPa'), ".1f"),
        "f'c Vigas/Losas (MPa)": format_value(pg.get('fc_losas_vigas_MPa'), ".1f"),
        "f'c Zapatas (MPa)": format_value(pg.get('fc_zapatas_MPa'), ".1f"),
        "f'y Acero Refuerzo (MPa)": format_value(pg.get('fy_MPa'), ".0f"),
        "Ec Concreto (aprox. MPa)": ec_concreto_str,
        "Es Acero (MPa)": "200000"
    }
    current_row_pt = _escribir_tabla_simple_v2(ws_portada, material_data_dict, current_row_pt, tf_pt, pks_pt, pvs_pt, "Parámetros de Materiales")

    sismicos_data_dict = {
        "Coeficiente Aa": format_value(pg.get('Aa'), ".2f"), 
        "Coeficiente Av": format_value(pg.get('Av'), ".2f"),
        "Tipo de Perfil de Suelo": pg.get('suelo_tipo', 'N/D'),
        "Coeficiente Fa": format_value(pg.get('Fa', 0.0), ".2f"), 
        "Coeficiente Fv": format_value(pg.get('Fv', 0.0), ".2f"),
        "Grupo de Uso": pg.get('grupo_uso', 'N/D'), 
        "Coeficiente de Importancia (I)": format_value(pg.get('I_coef', 0.0), ".2f"),
        "Sistema Estructural (R₀)": f"{pg.get('sistema_estructural_R0_desc', 'N/D')} (R₀={format_value(pg.get('R0', 0.0), '.1f')})",
        "Factor ΦA (Altura)": format_value(datos_memoria.get('phi_A_final', pg.get('phi_A_usado',1.0)), ".2f"),
        "Factor ΦP (Planta)": format_value(datos_memoria.get('phi_P_final', pg.get('phi_P_usado',1.0)), ".2f"),
        "Factor ΦE (Redundancia)": "1.00 (Asumido)",
        "Coeficiente R Final Usado": format_value(datos_memoria.get('R_final_usado_espectro', 0.0), ".2f")
    }
    current_row_pt = _escribir_tabla_simple_v2(ws_portada, sismicos_data_dict, current_row_pt, tf_pt, pks_pt, pvs_pt, "Parámetros Sísmicos (NSR-10)")

    # --- Hoja: Criterios y Cargas de Diseño ---
    ws_cargas = wb.create_sheet("Criterios y Cargas")
    tf_cc, hs_cc, ns_cc, pks_cc, pvs_cc = aplicar_estilos_reporte(ws_cargas)
    current_row_cc = 1
    info_cargas = datos_memoria.get("info_cargas_criterios", {})
    current_row_cc = _escribir_tabla_simple_v2(ws_cargas, {
        "Descripción del Proyecto": info_cargas.get("descripcion_proyecto_detallada","N/D"),
        "Normativa de Referencia": info_cargas.get("normativa_referencia","N/D"),
        "Software Utilizado": info_cargas.get("software_usado","N/D"),
        "Criterios DMO Aplicados": info_cargas.get("criterios_dmo_aplicados","N/D")
    }, current_row_cc, tf_cc, pks_cc, pvs_cc, "Criterios Generales de Diseño")

    if info_cargas.get("cargas_muertas_tipicas"):
        df_cm = pd.DataFrame(info_cargas["cargas_muertas_tipicas"])
        if not df_cm.empty:
            current_row_cc = _escribir_dataframe_a_hoja_v2(ws_cargas, df_cm, current_row_cc, tf_cc, hs_cc, ns_cc, "Cargas Muertas Típicas (kN/m² o kN/m)")
    if info_cargas.get("cargas_vivas_tipicas"):
        df_cv = pd.DataFrame(info_cargas["cargas_vivas_tipicas"])
        if not df_cv.empty:
            current_row_cc = _escribir_dataframe_a_hoja_v2(ws_cargas, df_cv, current_row_cc, tf_cc, hs_cc, ns_cc, "Cargas Vivas de Diseño (NSR-10 B.4)")

    # --- Hoja: Espectro Sísmico ---
    espectro_data_dict = datos_memoria.get("espectro_calculado_data")
    pg_esp = datos_memoria.get("parametros_globales", {})
    if espectro_data_dict and espectro_data_dict.get("T") is not None:
        ws_espectro = wb.create_sheet("Espectro Sísmico")
        tf_es, hs_es, ns_es, pks_es, pvs_es = aplicar_estilos_reporte(ws_espectro)
        current_row_es = 1
        info_p = espectro_data_dict.get("info_periodos", {})
        data_periodos_es = {
            "Tipo Espectro": espectro_data_dict.get("tipo", "N/D").capitalize(),
            "R usado": format_value(espectro_data_dict.get('R_usado', 0.0), ".2f"),
            "I usado": format_value(espectro_data_dict.get('I_usado', 0.0), ".2f"),
            "T₀ (s)": format_value(info_p.get('T0', 0.0), ".3f"),
            "Tᴄ (s)": format_value(info_p.get('TC', 0.0), ".3f"),
            "Tʟ (s)": format_value(info_p.get('TL_norma', 0.0), ".3f"),
        }
        current_row_es = _escribir_tabla_simple_v2(ws_espectro, data_periodos_es, current_row_es, tf_es, pks_es, pvs_es, "Parámetros del Espectro")
        
        df_espectro = pd.DataFrame({
            'Periodo T (s)': espectro_data_dict["T"],
            f'Sa ({espectro_data_dict.get("tipo","diseño")}) (g)': espectro_data_dict["Sa"]
        })
        current_row_es = _escribir_dataframe_a_hoja_v2(ws_espectro, df_espectro, current_row_es, tf_es, hs_es, ns_es, "Datos del Espectro")

        Aa_esp_val = pg_esp.get('Aa'); Fa_esp_val = pg_esp.get('Fa'); Av_esp_val = pg_esp.get('Av'); Fv_esp_val = pg_esp.get('Fv')
        Sds_val = "N/D"; Sd1_val = "N/D"
        try:
            if pd.notna(Aa_esp_val) and pd.notna(Fa_esp_val): Sds_val = 2.5 * float(Aa_esp_val) * float(Fa_esp_val)
            if pd.notna(Av_esp_val) and pd.notna(Fv_esp_val): Sd1_val = float(Av_esp_val) * float(Fv_esp_val)
        except (TypeError, ValueError): pass # Mantener "N/D" si la conversión falla

        data_sds_sd1 = {
            "S_DS (2.5*Aa*Fa) (g)": format_value(Sds_val, ".3f"),
            "S_D1 (Av*Fv) (g)": format_value(Sd1_val, ".3f")
        }
        current_row_es = _escribir_tabla_simple_v2(ws_espectro, data_sds_sd1, current_row_es, tf_es, pks_es, pvs_es, "Parámetros Espectrales Adicionales")
        
        path_img_esp = datos_memoria.get("path_imagen_espectro")
        img_anchor_col = openpyxl.utils.get_column_letter(df_espectro.shape[1] + 2) # Anclar a la derecha de la tabla
        if path_img_esp and os.path.exists(path_img_esp):
            try:
                img = OpenpyxlImage(path_img_esp)
                img.anchor = f"{img_anchor_col}2" 
                img.height = 300; img.width = 450
                ws_espectro.add_image(img)
            except Exception as e_img_add:
                ws_espectro.cell(row=current_row_es, column=1, value=f"Error al añadir imagen: {e_img_add}").font = Font(italic=True, color="FF0000")
        else:
            ws_espectro.cell(row=current_row_es, column=1, value="Gráfico del espectro no disponible.").font = Font(italic=True)

    # --- Hoja: Evaluación de Irregularidades ---
    ws_irr = wb.create_sheet("Irregularidades")
    tf_i, hs_i, ns_i, pks_i, pvs_i = aplicar_estilos_reporte(ws_irr)
    current_row_i = 1
    info_irr_rep_dict = datos_memoria.get("info_irregularidades", {})
    data_irr_summary = {
        "Factor ΦA (Altura) Usado": format_value(info_irr_rep_dict.get('phi_A_usado'), ".2f"),
        "Factor ΦP (Planta) Usado": format_value(info_irr_rep_dict.get('phi_P_usado'), ".2f"),
        "Comentarios Irregularidad en Planta": info_irr_rep_dict.get("evaluacion_planta", "No provista."),
        "Comentarios Irregularidad en Altura": info_irr_rep_dict.get("evaluacion_altura", "No provista.")
    }
    current_row_i = _escribir_tabla_simple_v2(ws_irr, data_irr_summary, current_row_i, tf_i, pks_i, pvs_i, "Resumen Evaluación de Irregularidades")

    # --- Hoja: Fuerza Horizontal Equivalente ---
    fhe_data_dict = datos_memoria.get("resultados_fhe")
    if fhe_data_dict and fhe_data_dict.get("df_Fx") is not None:
        ws_fhe = wb.create_sheet("Fuerza Horizontal Eq.")
        tf_fhe, hs_fhe, ns_fhe, pks_fhe, pvs_fhe = aplicar_estilos_reporte(ws_fhe)
        current_row_fhe = 1
        data_fhe_resumen = {
            "Peso Sísmico Total W (kN)": format_value(datos_memoria.get("peso_sismico_total_usado_para_fhe"),".2f"),
            "Periodo Ta (s)": format_value(datos_memoria.get('Ta_calculado_para_fhe'), ".3f"),
            "Sa(Ta) (g)": format_value(datos_memoria.get('Sa_Ta_usado_para_fhe'), ".4f"),
            "Cortante Basal Vs (kN)": format_value(fhe_data_dict.get('Vs_kN'), ".2f"),
            "Exponente k (distribución)": format_value(fhe_data_dict.get('k_dist'), ".3f")
        }
        current_row_fhe = _escribir_tabla_simple_v2(ws_fhe, data_fhe_resumen, current_row_fhe, tf_fhe, pks_fhe, pvs_fhe, "Resumen FHE")
        df_Fx_dist = fhe_data_dict["df_Fx"]
        current_row_fhe = _escribir_dataframe_a_hoja_v2(ws_fhe, df_Fx_dist, current_row_fhe, tf_fhe, hs_fhe, ns_fhe, "Distribución de Fuerzas Sísmicas por Nivel")

    # --- Hojas de Diseño de Elementos (Columnas, Vigas, Zapatas, etc.) ---
    element_sheets = {
        "Diseño Columnas": {
            "flex_key": "columnas_flexion_disenadas", 
            "cort_key": "columnas_cortante_disenadas",
            "flex_title": "Diseño Columnas (Flexo-compresión)",
            "cort_title": "Diseño Columnas (Cortante y Confinamiento DMO)",
            "flex_cols": { "ID Columna": "ID", "b (cm)": "b", "h (cm)": "h", "Rec. Libre (cm)": "Rec.", "f'c (MPa)": "f'c", "fy (MPa)": "fy", "Ø Barra Long. (mm)": "ØLong", "Ø Estribo (mm)": "ØEstr", "Nx Barras (cara b)": "Nx", "Ny Barras (cara h, interm.)": "Ny", "Nº Total Barras": "NºTotal", "As Total (cm²)": "As Total", "Cuantía (ρg)": "ρg", "φPn_max (kN)": "φPn_max", "Estado Diagrama": "Estado"},
            "cort_cols": { "ID Columna": "ID", "Vu Diseño (kN)": "Vu (kN)", "Ve Capacidad (kN)": "Ve (kN)", "Vc (kN)": "Vc (kN)", "Vs Req (kN)": "Vs_req (kN)", "Lo (cm)": "Lo (cm)", "Ø Estribo Usado (mm)": "ØEstribo", "s Confinado Final (mm)": "s conf (mm)", "s Central Final (mm)": "s central (mm)", "Estado Cortante": "Estado"}
        },
        "Diseño Vigas (DMO)": {
            "data_key": "vigas_disenadas", "title": "Resumen Diseño de Vigas (DMO)",
            "cols": { "ID Elemento": "ID Viga", "b (cm)": "b", "h (cm)": "h", "d (cm)": "d", "Ln (m)": "Ln", "Mu(-) Ext (kNm)": "Mu(-)Ext", "As(-) Ext (cm²)": "As(-)Ext", "ρ(-) Ext": "ρ(-)Ext", "Mu(+) Cen (kNm)": "Mu(+)Cen", "As(+) Cen (cm²)": "As(+)Cen", "ρ(+) Cen": "ρ(+)Cen", "Mu(-) Int (kNm)": "Mu(-)Int", "As(-) Int (cm²)": "As(-)Int", "ρ(-) Int": "ρ(-)Int", "Ve,max (kN)": "Ve,max", "Ø Estribo (mm)": "ØEstribo", "s confinado (cm)": "s conf.", "s central (cm)": "s central", "Lo (cm)": "Lo", "Estado Diseño": "Estado"},
            "max_col_width": 15
        },
        "Diseño Zapatas": {
            "data_key": "zapatas_disenadas", "title": "Resumen Diseño de Zapatas", # Clave de app.py para zapatas
            "cols": {"ID Zapata": "ID", "B (m)": "B", "L (m)": "L", "h (m)": "h", "d prom (m)": "d", "P_serv (kN)": "Pserv", "P_ult (kN)": "Pult", "Mux_ult_diseno (kNm)": "Mux,u", "Muy_ult_diseno (kNm)": "Muy,u", "q_max_serv (kPa)": "q_max,s", "q_min_serv (kPa)": "q_min,s", "q_adm (kPa)": "q_adm", "q_max_ult (kPa)": "q_max,u", "Cort_Uni_L_OK": "Cort.L OK?", "Cort_Uni_B_OK": "Cort.B OK?", "Punz_OK": "Punz.OK?", "As_L_cm2/m": "As,L (cm²/m)", "As_B_cm2/m": "As,B (cm²/m)"},
            "max_col_width": 15
        },
        "Diseño Losas Macizas": {
            "data_key": "losas_macizas_disenadas", "title": "Diseño de Losas Macizas Unidireccionales",
            "cols": { "ID Losa": "ID", "h (cm)": "h", "Mu (kNm/m)": "Mu", "d efectivo (cm)": "d", "As Ppal (cm²/m)": "As Ppal", "Ref. Ppal": "Ref.Ppal", "As Temp (cm²/m)": "As Temp", "Ref. Temp.": "Ref.Temp", "Verif. Espesor": "Nota Espesor"},
             "max_col_width": 25
        },
        "Diseño Losas Nervadas": {
            "data_key": "nervios_disenados", "title": "Diseño de Nervios de Losa Aligerada",
            "cols": { "ID Elemento": "ID Nervio", "h_total (cm)": "h", "bw (cm)": "bw", "hf (cm)": "hf", "L_libre (m)": "Ln", "S_nervios (m)": "S nervios", "Mu (kNm)": "Mu", "As_final (cm²)": "As", "Vu (kN)": "Vu", "Av/s (mm²/m)": "Av/s", "ØEstribo (mm)": "ØEstribo", "s_rec (mm)": "s estribo", "Estado Flexión": "Estado Flex.", "Estado Cortante": "Estado Cort."},
            "max_col_width": 15
        },
        "Diseño Escaleras": {
            "data_key": "escaleras_disenadas", "title": "Diseño de Tramos de Escalera",
            "cols": { "ID Tramo": "ID", "L_horiz (m)": "L_horiz", "h_garganta (cm)": "h garg.", "Mu (kNm/m)": "Mu", "d (cm)":"d", "As Ppal (cm²/m)": "As Ppal", "Ref. Ppal": "Ref.Ppal", "As Temp (cm²/m)": "As Temp", "Ref. Temp.": "Ref.Temp", "Verif. Espesor": "Nota Espesor"},
            "max_col_width": 20
        },
        "Verificación Deflexiones": {
            "data_key": "deflexiones_verificadas", "title": "Resumen Verificación de Deflexiones",
            "cols": { "ID Elemento": "ID Elem.", "Tipo Elemento": "Tipo", "Luz Libre Ln (cm)": "Ln (cm)", "Ie (mm⁴)": "Ie (mm⁴)", "Δ inst CM (mm)": "ΔCM_inst", "Δ inst CVtot (mm)": "ΔCV_inst", "Δ adic LP (mm)": "ΔLP_adic", "Δ verif susc (mm)": "Δverif_susc", "Límite CV (L/n)": "Límite CV", "Cumple CV": "Cumple CV?", "Límite Dif. (L/n)": "Límite Dif.", "Cumple Dif.": "Cumple Dif.?"},
            "max_col_width": 20
        }
    }

    for sheet_name, config in element_sheets.items():
        if sheet_name == "Diseño Columnas": # Caso especial para columnas con dos tablas
            flex_list = datos_memoria.get(config["flex_key"], [])
            cort_list = datos_memoria.get(config["cort_key"], [])
            if flex_list or cort_list:
                ws = wb.create_sheet(sheet_name)
                tf_s, hs_s, ns_s, _, _ = aplicar_estilos_reporte(ws)
                current_row_s = 1
                if flex_list:
                    df_flex = pd.DataFrame(flex_list)
                    df_flex_rep = df_flex[[k for k in config["flex_cols"] if k in df_flex.columns]].rename(columns=config["flex_cols"])
                    current_row_s = _escribir_dataframe_a_hoja_v2(ws, df_flex_rep, current_row_s, tf_s, hs_s, ns_s, config["flex_title"])
                if cort_list:
                    df_cort = pd.DataFrame(cort_list)
                    df_cort_rep = df_cort[[k for k in config["cort_cols"] if k in df_cort.columns]].rename(columns=config["cort_cols"])
                    _escribir_dataframe_a_hoja_v2(ws, df_cort_rep, current_row_s, tf_s, hs_s, ns_s, config["cort_title"])
        else:
            data_list = datos_memoria.get(config["data_key"], [])
            if data_list:
                ws = wb.create_sheet(sheet_name)
                tf_s, hs_s, ns_s, _, _ = aplicar_estilos_reporte(ws)
                df_data = pd.DataFrame(data_list)
                if not df_data.empty:
                    df_report = df_data[[k for k in config["cols"] if k in df_data.columns]].rename(columns=config["cols"])
                    _escribir_dataframe_a_hoja_v2(ws, df_report, 1, tf_s, hs_s, ns_s, config["title"], max_col_width=config.get("max_col_width",30))


    # --- Hoja: Combinaciones de Carga ---
    comb_data_dict = datos_memoria.get("combinaciones_usadas")
    if comb_data_dict:
        ws_comb = wb.create_sheet("Combinaciones Carga")
        tf_c, hs_c, ns_c, _, _ = aplicar_estilos_reporte(ws_comb) # Solo 3 estilos
        current_row_c = 1
        if "servicio" in comb_data_dict:
            serv_list = [{"Combinación": nom, **fac} for nom, fac in comb_data_dict["servicio"]]
            df_serv = pd.DataFrame(serv_list).fillna(0)
            if not df_serv.empty:
                current_row_c = _escribir_dataframe_a_hoja_v2(ws_comb, df_serv, current_row_c, tf_c, hs_c, ns_c, "Combinaciones de Servicio")
        if "ultimas" in comb_data_dict:
            ult_list = [{"Combinación": nom, **fac} for nom, fac in comb_data_dict["ultimas"]]
            df_ult = pd.DataFrame(ult_list).fillna(0)
            if not df_ult.empty:
                _escribir_dataframe_a_hoja_v2(ws_comb, df_ult, current_row_c, tf_c, hs_c, ns_c, "Combinaciones Últimas (Mayoradas)")

    # --- Guardar el libro de trabajo ---
    try:
        wb.save(nombre_archivo)
        path_img_esp_final = datos_memoria.get("path_imagen_espectro")
        if path_img_esp_final and os.path.exists(path_img_esp_final) and "temp" in path_img_esp_final :
            try: os.remove(path_img_esp_final)
            except Exception: pass # No fallar si no se puede borrar
        return f"Memoria de cálculo generada exitosamente: {nombre_archivo}"
    except PermissionError:
        return f"Error de Permiso: No se pudo guardar '{nombre_archivo}'. Asegúrate de que no esté abierto o protegido."
    except Exception as e:
        return f"Error al guardar el archivo Excel: {e}"